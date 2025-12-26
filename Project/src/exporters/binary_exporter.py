# exporters/binary_exporter.py

import os
import struct
from openpyxl import load_workbook
from .base import BaseExporter
from schema.types import BasicType, EnumType, ArrayType, CustomType

class BinaryExporter(BaseExporter):
    """
    二进制导出器示例：
    - 每个字段按类型打包（int/float/string等）
    - 支持基本类型和枚举
    - 自定义类型暂时当基础类型写入（或扩展逻辑）
    """

    file_ext = "bin"

    def export_data(self, file_path, models, enums):
        """
        解析 Excel DataTables，返回 dict: {model_name: 数据列表}
        """
        wb = load_workbook(file_path, data_only=True)
        result = {}

        for model in models:
            sheet_name = model.name
            if sheet_name not in wb.sheetnames:
                print(f"跳过 {sheet_name}, sheet 不存在")
                continue
            ws = wb[sheet_name]

            data_list = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue

                obj = {}
                for field, value in zip(model.fields, row):
                    if isinstance(field.type, BasicType):
                        obj[field.name] = value
                    elif isinstance(field.type, EnumType):
                        if isinstance(value, int):
                            obj[field.name] = value
                        elif isinstance(value, str):
                            obj[field.name] = field.type.members.get(value, 0)
                        else:
                            obj[field.name] = 0
                    elif isinstance(field.type, CustomType):
                        obj[field.name] = value
                    elif isinstance(field.type, ArrayType):
                        if isinstance(value, str):
                            obj[field.name] = [e.strip() for e in value.split(",")]
                        else:
                            obj[field.name] = []
                data_list.append(obj)
            result[model.name] = data_list
        return result

    def write_file(self, data_dict, output_dir):
        """
        将数据写入二进制文件，每个 model 一个文件
        """
        os.makedirs(output_dir, exist_ok=True)
        for model_name, data_list in data_dict.items():
            out_file = os.path.join(output_dir, f"{model_name}.{self.file_ext}")
            with open(out_file, "wb") as f:
                for obj in data_list:
                    for key, value in obj.items():
                        # 简单示例，只处理 int/float/string
                        if isinstance(value, int):
                            f.write(struct.pack("<i", value))
                        elif isinstance(value, float):
                            f.write(struct.pack("<f", value))
                        elif isinstance(value, str):
                            encoded = value.encode("utf-8")
                            f.write(struct.pack("<I", len(encoded)))
                            f.write(encoded)
                        elif isinstance(value, list):
                            # 列表长度 + 元素字符串
                            f.write(struct.pack("<I", len(value)))
                            for item in value:
                                item_str = str(item).encode("utf-8")
                                f.write(struct.pack("<I", len(item_str)))
                                f.write(item_str)
                        else:
                            # 其他类型写空
                            f.write(struct.pack("<I", 0))
            print(f"导出 DataTable {model_name} 到 {out_file}")
