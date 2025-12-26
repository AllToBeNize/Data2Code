import openpyxl
from openpyxl.utils import range_boundaries

from schema.types import (
    BasicType,
    ArrayType,
    EnumType,
    CustomType,
)
from schema.field import FieldDef
from schema.model import ModelDef


# =========================
# 基础类型表
# =========================

BASIC_TYPES = {
    "int": BasicType("int"),
    "float": BasicType("float"),
    "string": BasicType("string"),
    "bool": BasicType("bool"),
}


# =========================
# Excel Table 扫描
# =========================

def scan_tables(ws):
    """
    扫描 worksheet 中的所有 Excel「表格对象」
    返回：
        [
            (table_name, headers, data_rows),
            ...
        ]
    """
    tables = []

    for table in ws.tables.values():
        table_name = table.name
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)

        headers = [
            ws.cell(row=min_row, column=col).value
            for col in range(min_col, max_col + 1)
        ]

        data_rows = []
        for row in range(min_row + 1, max_row + 1):
            row_values = [
                ws.cell(row=row, column=col).value
                for col in range(min_col, max_col + 1)
            ]
            # 跳过空行
            if any(v is not None for v in row_values):
                data_rows.append(row_values)

        tables.append((table_name, headers, data_rows))

    return tables


# =========================
# Enums 解析
# =========================

def parse_enums_sheet(file_path, sheet_name="Enums"):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    tables = scan_tables(ws)

    enum_map = {}

    for table_name, headers, rows in tables:
        members = {}
        underlying = "int"  # 默认 underlying

        for row in rows:
            name = row[0]
            if name is None:
                continue

            # 读取 underlying 声明
            if isinstance(name, str) and name.startswith("__"):
                if name == "__underlying" and len(row) > 1 and row[1]:
                    underlying = str(row[1]).strip()
                continue

            # 枚举成员
            if len(row) > 1 and row[1] is not None:
                value = int(row[1])
            else:
                value = len(members)

            members[name] = value

        enum_map[table_name] = EnumType(
            name=table_name,
            members=members,
            underlying=underlying,
        )

    return enum_map


# =========================
# 类型解析
# =========================

def parse_type(type_str, model_map, enum_map):
    """
    解析字段类型字符串
    支持：
        int
        float
        string
        bool
        SkillConfig
        SkillType
        int[]
        SkillConfig[]
    """
    if not type_str:
        raise ValueError("字段类型不能为空")

    type_str = str(type_str).strip()

    # 数组
    if type_str.endswith("[]"):
        elem = parse_type(type_str[:-2], model_map, enum_map)
        return ArrayType(elem)

    key = type_str.lower()

    if key in BASIC_TYPES:
        return BASIC_TYPES[key]

    if type_str in enum_map:
        return enum_map[type_str]

    if type_str in model_map:
        return CustomType(type_str)

    raise ValueError(f"未知类型: {type_str}")


# =========================
# Models 解析
# =========================

def parse_models_sheet(file_path, enum_map, sheet_name="Models"):
    """
    Models Sheet 规则：
    每个表：
        FieldName | Type | Comment | PrimaryKey
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    tables = scan_tables(ws)

    # 先注册所有 Model（解决互相引用）
    model_map = {}
    table_rows = {}

    for table_name, headers, rows in tables:
        model_map[table_name] = ModelDef(
            name=table_name,
            fields=[]
        )
        table_rows[table_name] = rows

    # 第二遍解析字段
    for model_name, rows in table_rows.items():
        model = model_map[model_name]

        for row in rows:
            if len(row) < 2:
                continue

            field_name = row[0]
            type_name = row[1]
            comment = row[2] if len(row) > 2 else ""
            primary_raw = row[3] if len(row) > 3 else False

            if not field_name or not type_name:
                continue

            is_primary = False
            if isinstance(primary_raw, bool):
                is_primary = primary_raw
            elif isinstance(primary_raw, (int, float)):
                is_primary = primary_raw != 0
            elif isinstance(primary_raw, str):
                is_primary = primary_raw.strip().lower() in ("true", "1", "yes")

            field_type = parse_type(type_name, model_map, enum_map)

            model.fields.append(
                FieldDef(
                    name=field_name,
                    type=field_type,
                    comment=comment or "",
                    is_primary=is_primary,
                )
            )

    return list(model_map.values())


# =========================
# Schema 校验
# =========================

def validate_schema(file_path, models, enums, sheet_models="Models", sheet_enums="Enums"):
    """
    校验解析后的 schema
    Args:
        file_path: Excel 文件路径
        models: parse_models_sheet 返回的 List[ModelDef]
        enums: parse_enums_sheet 返回的 Dict[str, EnumType]
    """
    errors = []

    # ------------------------
    # 1. 检查 Model 重名
    # ------------------------
    model_names = set()
    for model in models:
        if model.name in model_names:
            errors.append(f"[{file_path}][{sheet_models}] 重复 Model 名: {model.name}")
        model_names.add(model.name)

    # ------------------------
    # 2. 检查 Enum 重名
    # ------------------------
    enum_names = set()
    for enum_name in enums.keys():
        if enum_name in enum_names:
            errors.append(f"[{file_path}][{sheet_enums}] 重复 Enum 名: {enum_name}")
        enum_names.add(enum_name)

    # ------------------------
    # 3. 检查 Model 内字段重复 & 主键重复
    # ------------------------
    for model in models:
        field_names = set()
        primary_count = 0
        for field in model.fields:
            # 重复字段
            if field.name in field_names:
                errors.append(f"[{file_path}][{sheet_models}][{model.name}] 重复字段: {field.name}")
            field_names.add(field.name)

            # 主键数量
            if field.is_primary:
                primary_count += 1

        if primary_count > 1:
            errors.append(f"[{file_path}][{sheet_models}][{model.name}] 主键重复 (多个 PrimaryKey)")

    # ------------------------
    # 4. 输出结果
    # ------------------------
    if errors:
        for err in errors:
            print("Schema 校验错误:", err)
        raise ValueError(f"{file_path} 校验失败，共 {len(errors)} 个错误")
    else:
        print(f"{file_path} 校验通过，没有发现重复主键或重名类型。")

