# exporters/json_exporter.py
import os
import json
from openpyxl import load_workbook
from .base import BaseDataExporter
from schema.types import BasicType, EnumType, ArrayType, CustomType

class JSONExporter(BaseDataExporter):
    file_ext = "json"

    def export_data(self, file_path, models, enums):
        wb = load_workbook(file_path, data_only=True)
        data_dict = {}

        for model in models:
            if model.name not in wb.sheetnames:
                continue
            ws = wb[model.name]
            data_list = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                obj = {}
                for field, value in zip(model.fields, row):
                    obj[field.name] = self._convert_value(field.type, value)
                data_list.append(obj)
            data_dict[model.name] = data_list
        return data_dict

    def _convert_value(self, field_type, value):
        if isinstance(field_type, BasicType):
            return value
        elif isinstance(field_type, EnumType):
            if isinstance(value, int):
                return value
            elif isinstance(value, str):
                return field_type.members.get(value, 0)
            else:
                return 0
        elif isinstance(field_type, CustomType):
            return value
        elif isinstance(field_type, ArrayType):
            if not value:
                return []
            if isinstance(field_type.element_type, (CustomType, ArrayType)):
                raise TypeError(f"数组字段 {field_type.name} 不支持自定义类型或嵌套数组")
            if isinstance(value, str):
                elems = [e.strip() for e in value.split(",")]
            else:
                elems = value
            converted = []
            for e in elems:
                if isinstance(field_type.element_type, BasicType):
                    if field_type.element_type.name == "int":
                        converted.append(int(e))
                    elif field_type.element_type.name == "float":
                        converted.append(float(e))
                    elif field_type.element_type.name == "bool":
                        converted.append(e.lower() in ("1", "true", "yes") if isinstance(e, str) else bool(e))
                    else:
                        converted.append(str(e))
                elif isinstance(field_type.element_type, EnumType):
                    if isinstance(e, int):
                        converted.append(e)
                    else:
                        converted.append(field_type.element_type.members.get(str(e), 0))
            return converted
        else:
            return value

    def write_file(self, data_dict, output_dir):
        os.makedirs(output_dir, exist_ok=True)
        mapping = {}
        for model_name, data_list in data_dict.items():
            out_file_name = f"DT_{model_name}.{self.file_ext}"
            out_file = os.path.join(output_dir, out_file_name)
            with open(out_file, "w", encoding="utf-8") as f:
                json.dump(data_list, f, ensure_ascii=False, indent=4)
            mapping[model_name] = out_file_name
            print(f"导出 DataTable {model_name} 到 {out_file}")
        # 生成 mapping.json
        mapping_file = os.path.join(output_dir, "_mapping.json")
        with open(mapping_file, "w", encoding="utf-8") as f:
            json.dump(mapping, f, ensure_ascii=False, indent=4)
        print(f"生成映射文件 {mapping_file}")
